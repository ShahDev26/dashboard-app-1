# -*- coding: utf-8 -*-
"""
Build script for the deployable Vercel dashboard.

Reads Admin_MasterData_TestPlan.xlsx from ../../testing and emits
public/index.html with all TC data embedded. The dashboard fetches live
status overrides from /api/status on load and POSTs status changes back.

Run:
    cd dashboard-app
    python scripts/build.py
"""
import io, json, re, sys
from datetime import datetime
from pathlib import Path

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

# On Vercel (or any environment without openpyxl), exit cleanly.
# The dashboard is pre-built locally and public/index.html is committed.
try:
    import openpyxl
except ImportError:
    print('openpyxl not installed — skipping rebuild (public/index.html is pre-built).')
    sys.exit(0)

# Also exit cleanly if the source workbook is not bundled with the deploy.
if not (Path(__file__).resolve().parent.parent.parent / 'testing' / 'Admin_MasterData_TestPlan.xlsx').exists():
    print('Source workbook not found in this checkout — skipping rebuild (public/index.html is pre-built).')
    sys.exit(0)

HERE = Path(__file__).resolve().parent
APP = HERE.parent
import os
# Allow MEL_XLSX env var to override (useful when Excel has the workbook locked).
XLSX = Path(os.environ.get('MEL_XLSX', str(APP.parent / 'testing' / 'Admin_MasterData_TestPlan.xlsx')))
OUT = APP / 'public' / 'index.html'

MODULES = [
    ('Core', 'Authentication',            'auth',     'Authentication',          '🔐', 'tc'),
    ('Core', 'AM-01 Users',               'users',    'Users',                   '👥', 'tc'),
    ('Master Data', 'AM-02 GSC Centres',          'centres',   'GSC Centres',             '🏢', 'tc'),
    ('Master Data', 'AM-04 State',                'states',    'States',                  '🗺️', 'tc'),
    ('Master Data', 'AM-05 District',             'districts', 'Districts',               '📍', 'tc'),
    ('Master Data', 'AM-06 Block-Taluk',          'blocks',    'Blocks / Taluks',         '🏘️', 'tc'),
    ('Master Data', 'AM-07 Village',              'villages',  'Villages',                '🌾', 'tc'),
    ('Master Data', 'AM-08 Pincode',              'pincodes',  'Pincodes',                '📮', 'tc'),
    ('Master Data', 'AM-03 Disability',           'disability','Disability Types',        '♿', 'tc'),
    ('Master Data', 'AM-09 Course Type',          'courseType','Course Types',            '📚', 'tc'),
    ('Master Data', 'AM-10 Impl Partners',        'iPartners', 'Implementation Partners', '🤝', 'tc'),
    ('Master Data', 'AM-11 Course Partners',      'cPartners', 'Course Partners',         '🎓', 'tc'),
    ('Master Data', 'AM-12 Financial Partners',   'fPartners', 'Financial Partners',      '💰', 'tc'),
    ('Quality', 'WCAG Accessibility',             'a11y',     'Accessibility (WCAG)',     '⌨️', 'wcag'),
    ('Quality', 'Defect Log',                     'defects',  'Defect Log',               '🐛', 'defects'),
    ('Quality', 'Resolutions',                    'resol',    'Resolutions',              '✅', 'resol'),
]

ROLE_PATTERNS = [
    ('Admin',      re.compile(r'\b(?:Admin|admin)\b')),
    ('State Head', re.compile(r'\b(?:State Head|SH|state-head)\b')),
    ('GSCO',       re.compile(r'\b(?:GSCO|gsco|G S C O)\b')),
]

def detect_roles(*texts):
    blob = ' '.join(t for t in texts if t)
    found = [r for r, p in ROLE_PATTERNS if p.search(blob)]
    if not found:
        found = ['Any']
    return found

def safe(cell):
    v = cell.value if hasattr(cell, 'value') else cell
    return '' if v is None else str(v)

def parse_touchpoints(raw):
    """Parse the Touchpoints cell into a list of {role, screen, action} dicts.

    Each line in the cell is one touchpoint formatted as `Role : Screen : Action`.
    Lines that don't have at least one colon are accepted as a free-text fallback
    (placed in `action`, with role/screen empty). Blank lines and empty cells
    are skipped.
    """
    out = []
    for line in (raw or '').splitlines():
        line = line.strip().lstrip('-').lstrip('*').strip()
        if not line:
            continue
        parts = [p.strip() for p in line.split(':', 2)]
        if len(parts) == 3:
            out.append({'role': parts[0], 'screen': parts[1], 'action': parts[2]})
        elif len(parts) == 2:
            out.append({'role': parts[0], 'screen': '', 'action': parts[1]})
        else:
            out.append({'role': '', 'screen': '', 'action': parts[0]})
    return out


wb = openpyxl.load_workbook(XLSX, data_only=False)

modules_out = []
total_tcs = 0

for category, sheet_name, mod_id, mod_name, icon, layout in MODULES:
    if sheet_name not in wb.sheetnames:
        continue
    ws = wb[sheet_name]
    tcs = []

    if layout == 'tc':
        for r in range(2, ws.max_row + 1):
            tcid = safe(ws.cell(row=r, column=1))
            if not tcid: continue
            scenario = safe(ws.cell(row=r, column=2))
            ttype = safe(ws.cell(row=r, column=3)) or 'Functional'
            pre = safe(ws.cell(row=r, column=4))
            steps = safe(ws.cell(row=r, column=5))
            expected = safe(ws.cell(row=r, column=6))
            status = safe(ws.cell(row=r, column=8)) or 'Not run'
            severity = safe(ws.cell(row=r, column=9)) or 'Medium'
            br = safe(ws.cell(row=r, column=10))
            notes = safe(ws.cell(row=r, column=11))
            touchpoints = parse_touchpoints(safe(ws.cell(row=r, column=12)))
            tcs.append({
                'id': tcid, 'scenario': scenario, 'type': ttype,
                'pre': pre, 'steps': steps, 'expected': expected,
                'status': status, 'severity': severity, 'br': br,
                'notes': notes, 'roles': detect_roles(scenario, pre, steps, expected),
                'touchpoints': touchpoints,
            })
    elif layout == 'wcag':
        for r in range(5, ws.max_row + 1):
            tcid = safe(ws.cell(row=r, column=1))
            if not tcid: continue
            extras = []
            wcag_sc = safe(ws.cell(row=r, column=2));  level = safe(ws.cell(row=r, column=3))
            principle = safe(ws.cell(row=r, column=4)); surface = safe(ws.cell(row=r, column=6))
            viewport = safe(ws.cell(row=r, column=12))
            if wcag_sc: extras.append(f'WCAG SC: {wcag_sc} ({level})')
            if principle: extras.append(f'Principle: {principle}')
            if surface: extras.append(f'Surface: {surface}')
            if viewport: extras.append(f'Viewport: {viewport}')
            scenario = safe(ws.cell(row=r, column=5))
            pre = safe(ws.cell(row=r, column=7))
            steps = safe(ws.cell(row=r, column=8))
            expected = safe(ws.cell(row=r, column=9))
            touchpoints = parse_touchpoints(safe(ws.cell(row=r, column=15)))
            tcs.append({
                'id': tcid, 'scenario': scenario, 'type': 'Accessibility',
                'pre': pre, 'steps': steps, 'expected': expected,
                'status': safe(ws.cell(row=r, column=13)) or 'Not run',
                'severity': safe(ws.cell(row=r, column=11)) or 'Serious',
                'br': ' · '.join(extras),
                'notes': safe(ws.cell(row=r, column=14)),
                'roles': detect_roles(scenario, pre, steps),
                'touchpoints': touchpoints,
            })
    elif layout == 'defects':
        for r in range(4, ws.max_row + 1):
            tcid = safe(ws.cell(row=r, column=1))
            if not tcid: continue
            extras = []
            linked = safe(ws.cell(row=r, column=2))
            mod = safe(ws.cell(row=r, column=3))
            priority = safe(ws.cell(row=r, column=8))
            reporter = safe(ws.cell(row=r, column=9))
            date = safe(ws.cell(row=r, column=10))
            if linked: extras.append(f'Linked TC: {linked}')
            if mod: extras.append(f'Module: {mod}')
            if priority: extras.append(f'Priority: {priority}')
            if reporter: extras.append(f'Reported by: {reporter}')
            if date: extras.append(f'Date: {date}')
            title = safe(ws.cell(row=r, column=4))
            desc = safe(ws.cell(row=r, column=5))
            steps = safe(ws.cell(row=r, column=6))
            touchpoints = parse_touchpoints(safe(ws.cell(row=r, column=12)))
            tcs.append({
                'id': tcid, 'scenario': title, 'type': 'Bug',
                'pre': desc, 'steps': steps, 'expected': '',
                'status': safe(ws.cell(row=r, column=11)) or 'Open',
                'severity': safe(ws.cell(row=r, column=7)) or 'Medium',
                'br': ' · '.join(extras), 'notes': '',
                'roles': detect_roles(title, desc, steps),
                'touchpoints': touchpoints,
            })
    elif layout == 'resol':
        for r in range(4, ws.max_row + 1):
            tcid = safe(ws.cell(row=r, column=1))
            if not tcid: continue
            extras = []
            source = safe(ws.cell(row=r, column=2))
            linked = safe(ws.cell(row=r, column=7))
            if source: extras.append(f'Source: {source}')
            if linked: extras.append(f'Linked TC: {linked}')
            title = safe(ws.cell(row=r, column=3))
            desc = safe(ws.cell(row=r, column=4))
            verify = safe(ws.cell(row=r, column=6))
            touchpoints = parse_touchpoints(safe(ws.cell(row=r, column=9)))
            tcs.append({
                'id': tcid, 'scenario': title, 'type': 'Verification',
                'pre': desc, 'steps': verify, 'expected': '',
                'status': safe(ws.cell(row=r, column=8)) or 'Open',
                'severity': safe(ws.cell(row=r, column=5)) or 'Medium',
                'br': ' · '.join(extras), 'notes': '',
                'roles': detect_roles(title, desc, verify),
                'touchpoints': touchpoints,
            })

    if not tcs: continue
    modules_out.append({
        'category': category, 'id': mod_id, 'name': mod_name,
        'icon': icon, 'sheet': sheet_name, 'tcs': tcs,
    })
    total_tcs += len(tcs)


DATA = {
    'generatedAt': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
    'sourceFile': XLSX.name,
    'totalTcs': total_tcs,
    'modules': modules_out,
}

DATA_JSON = json.dumps(DATA, ensure_ascii=False).replace('</', '<\\/')

HTML = r"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8" />
<meta name="viewport" content="width=device-width,initial-scale=1" />
<title>MEL Test Dashboard</title>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
<link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&family=JetBrains+Mono:wght@400;500;600&display=swap" rel="stylesheet">
<style>
  :root {
    --bg:#f6f7fb; --surface:#ffffff; --surface-2:#f9fafc; --surface-hi:#eef0f6;
    --border:#e2e6ef; --border-hi:#cbd2de;
    --text:#0f172a; --text-2:#475569; --text-3:#94a3b8;
    --primary:#6366f1; --primary-hi:#4f46e5; --primary-soft:#eef2ff;
    --violet:#8b5cf6; --violet-soft:#f3eaff;
    --emerald:#10b981; --emerald-soft:#ecfdf5;
    --rose:#f43f5e; --rose-soft:#ffe4e6;
    --amber:#f59e0b; --amber-soft:#fef3c7;
    --orange:#fb923c; --orange-soft:#fff0e2;
    --red:#ef4444; --red-soft:#fee2e2;
    --slate:#64748b; --slate-soft:#f1f5f9;
    --shadow-sm:0 1px 2px rgba(15,23,42,0.04),0 1px 3px rgba(15,23,42,0.06);
    --shadow:0 4px 12px rgba(15,23,42,0.06),0 2px 4px rgba(15,23,42,0.04);
    --shadow-lg:0 16px 48px rgba(15,23,42,0.10),0 4px 12px rgba(15,23,42,0.06);
    --topbar-h:64px; --sidebar-w:272px;
  }
  *{box-sizing:border-box}html,body{margin:0;padding:0}
  body{background:var(--bg);color:var(--text);font-family:'Inter',-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;min-height:100vh;line-height:1.5;-webkit-font-smoothing:antialiased}
  .topbar{position:sticky;top:0;z-index:60;height:var(--topbar-h);background:var(--surface);border-bottom:1px solid var(--border);display:flex;align-items:center;padding:0 24px;gap:24px}
  .logo{display:flex;align-items:center;gap:10px;font-weight:800;font-size:17px;color:var(--text);cursor:pointer;letter-spacing:-0.02em}
  .logo-mark{width:32px;height:32px;border-radius:9px;background:linear-gradient(135deg,var(--primary),var(--violet));display:grid;place-items:center;color:#fff;font-size:18px;box-shadow:0 4px 10px rgba(99,102,241,0.35)}
  .logo-sub{color:var(--text-3);font-weight:500;font-size:13px}
  .topnav{display:flex;align-items:center;gap:6px;margin-left:8px}
  .role-chip{display:inline-flex;align-items:center;gap:7px;padding:7px 14px;background:var(--surface);border:1px solid var(--border);border-radius:8px;color:var(--text-2);font-size:13px;font-weight:500;cursor:pointer;user-select:none;transition:100ms}
  .role-chip:hover{background:var(--surface-hi);color:var(--text)}
  .role-chip .dot{width:6px;height:6px;border-radius:50%;background:var(--text-3)}
  .role-chip[data-role="Admin"] .dot{background:var(--primary)}
  .role-chip[data-role="State Head"] .dot{background:var(--violet)}
  .role-chip[data-role="GSCO"] .dot{background:var(--emerald)}
  .role-chip.on{background:var(--primary-soft);color:var(--primary-hi);border-color:var(--primary);font-weight:600}
  .role-chip.on[data-role="State Head"]{background:var(--violet-soft);color:var(--violet);border-color:var(--violet)}
  .role-chip.on[data-role="GSCO"]{background:var(--emerald-soft);color:var(--emerald);border-color:var(--emerald)}
  .top-stats{margin-left:auto;display:flex;align-items:center;gap:8px;font-family:'JetBrains Mono',monospace;font-size:12px;color:var(--text-2)}
  .stat-pill{padding:5px 11px;background:var(--surface-2);border:1px solid var(--border);border-radius:7px}
  .stat-pill b{color:var(--text);font-weight:700;margin-right:3px}
  .stat-pill.pass b{color:var(--emerald)}
  .stat-pill.fail b{color:var(--red)}
  .stat-pill.hold b{color:var(--amber)}
  .stat-pill.idle b{color:var(--slate)}
  .top-meta{color:var(--text-3);font-size:11px}
  .logout-btn{padding:7px 12px;background:var(--surface);border:1px solid var(--border);border-radius:8px;color:var(--text-2);font-family:inherit;font-size:12px;font-weight:500;cursor:pointer;transition:100ms}
  .logout-btn:hover{background:var(--surface-hi);color:var(--text);border-color:var(--border-hi)}
  .menu-btn{display:none;background:var(--surface);border:1px solid var(--border);border-radius:8px;width:36px;height:36px;align-items:center;justify-content:center;cursor:pointer;color:var(--text-2);font-size:18px}
  .shell{display:grid;grid-template-columns:var(--sidebar-w) 1fr;min-height:calc(100vh - var(--topbar-h))}
  .sidebar{position:sticky;top:var(--topbar-h);height:calc(100vh - var(--topbar-h));overflow-y:auto;background:var(--surface);border-right:1px solid var(--border);padding:18px 14px}
  .sidebar .home-link{display:flex;align-items:center;gap:10px;padding:9px 12px;border-radius:8px;color:var(--text-2);cursor:pointer;font-weight:600;font-size:13.5px;margin-bottom:14px}
  .sidebar .home-link:hover{background:var(--surface-hi);color:var(--text)}
  .sidebar .home-link.active{background:var(--primary-soft);color:var(--primary-hi)}
  .sidebar .category{margin-top:14px;padding:0 12px;font-size:11px;font-weight:700;color:var(--text-3);text-transform:uppercase;letter-spacing:0.10em}
  .sidebar .module-link{display:flex;align-items:center;gap:10px;padding:8px 12px;margin-top:2px;border-radius:8px;color:var(--text-2);font-size:13.5px;font-weight:500;cursor:pointer;transition:80ms}
  .sidebar .module-link:hover{background:var(--surface-hi);color:var(--text)}
  .sidebar .module-link.active{background:var(--primary-soft);color:var(--primary-hi);font-weight:600}
  .sidebar .module-link .ic{font-size:16px}
  .sidebar .module-link .count{margin-left:auto;font-family:'JetBrains Mono',monospace;font-size:11px;color:var(--text-3);background:var(--surface-hi);padding:2px 8px;border-radius:999px;min-width:30px;text-align:center}
  .sidebar .module-link.active .count{background:rgba(99,102,241,0.18);color:var(--primary-hi)}
  main{padding:26px 32px 80px;max-width:1400px}
  .hero{background:linear-gradient(135deg,#fff,#f5f7ff);border:1px solid var(--border);border-radius:16px;padding:28px 32px;margin-bottom:26px;position:relative;overflow:hidden}
  .hero::after{content:'';position:absolute;right:-60px;top:-60px;width:240px;height:240px;background:radial-gradient(circle,rgba(99,102,241,0.15),transparent 70%)}
  .hero h1{margin:0 0 6px;font-size:28px;font-weight:800;letter-spacing:-0.02em;color:var(--text)}
  .hero p{margin:0;color:var(--text-2);font-size:14px}
  .hero .role-banner{margin-top:14px;display:inline-flex;align-items:center;gap:8px;padding:6px 12px;background:var(--surface);border:1px solid var(--border);border-radius:999px;font-size:12px;color:var(--text-2)}
  .hero .role-banner b{color:var(--text);font-weight:600}
  .cat-block{margin-bottom:32px}
  .cat-title{font-size:12px;font-weight:700;color:var(--text-3);margin-bottom:12px;text-transform:uppercase;letter-spacing:0.1em}
  .module-grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(240px,1fr));gap:14px}
  .module-card{padding:18px;background:var(--surface);border:1px solid var(--border);border-radius:12px;cursor:pointer;transition:120ms}
  .module-card:hover{transform:translateY(-2px);border-color:var(--primary);box-shadow:var(--shadow)}
  .module-card .ic{font-size:28px;display:block;margin-bottom:10px}
  .module-card .name{font-weight:700;font-size:15px;color:var(--text)}
  .module-card .meta{font-size:12px;color:var(--text-2);margin-top:6px}
  .module-card .progress{height:4px;border-radius:999px;background:var(--surface-hi);margin-top:14px;overflow:hidden}
  .module-card .progress>div{height:100%;background:linear-gradient(90deg,var(--primary),var(--violet));transition:width 400ms ease}
  .crumb{font-size:12px;color:var(--text-3);margin-bottom:6px;font-family:'JetBrains Mono',monospace}
  .crumb a{color:var(--primary);text-decoration:none;cursor:pointer}
  .crumb a:hover{text-decoration:underline}
  .page-head{display:flex;align-items:center;gap:16px;margin-bottom:22px}
  .page-head .ic{width:56px;height:56px;border-radius:14px;background:var(--primary-soft);display:grid;place-items:center;font-size:28px}
  .page-head h1{margin:0;font-size:24px;font-weight:800;letter-spacing:-0.015em;color:var(--text)}
  .page-head .sub{font-size:13px;color:var(--text-2)}
  .filter-bar{background:var(--surface);border:1px solid var(--border);border-radius:12px;padding:14px 16px;margin-bottom:22px;display:flex;flex-wrap:wrap;gap:12px;align-items:center}
  .fselect{padding:8px 32px 8px 12px;border:1px solid var(--border);border-radius:8px;background-color:var(--surface);background-image:url("data:image/svg+xml;charset=utf-8,%3Csvg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 20 20' fill='%2364748b'%3E%3Cpath fill-rule='evenodd' d='M5.23 7.21a.75.75 0 011.06.02L10 11.06l3.71-3.83a.75.75 0 111.08 1.04l-4.25 4.39a.75.75 0 01-1.08 0L5.21 8.27a.75.75 0 01.02-1.06z' clip-rule='evenodd'/%3E%3C/svg%3E");background-repeat:no-repeat;background-position:right 10px center;background-size:14px;font-family:'Inter',sans-serif;font-size:13px;color:var(--text);cursor:pointer;appearance:none;-webkit-appearance:none;-moz-appearance:none;min-width:150px;transition:100ms}
  .fselect:hover{border-color:var(--border-hi);background-color:var(--surface-hi)}
  .fselect:focus{outline:none;border-color:var(--primary);box-shadow:0 0 0 3px rgba(99,102,241,0.15)}
  .fselect.active{border-color:var(--primary);background-color:var(--primary-soft);color:var(--primary-hi);font-weight:600}
  .search{flex:1;min-width:200px;padding:7px 12px 7px 32px;border:1px solid var(--border);border-radius:8px;background:var(--surface) url("data:image/svg+xml;charset=utf-8,%3Csvg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 20 20' fill='%2394a3b8'%3E%3Cpath fill-rule='evenodd' d='M9 3.5a5.5 5.5 0 100 11 5.5 5.5 0 000-11zM2 9a7 7 0 1112.452 4.391l3.328 3.329a.75.75 0 11-1.06 1.06l-3.329-3.328A7 7 0 012 9z' clip-rule='evenodd'/%3E%3C/svg%3E") no-repeat 10px center;background-size:14px;font-family:'Inter',sans-serif;font-size:13px;color:var(--text);outline:none}
  .search:focus{border-color:var(--primary);box-shadow:0 0 0 3px rgba(99,102,241,0.15)}
  .btn{padding:7px 14px;background:var(--surface);border:1px solid var(--border);border-radius:8px;color:var(--text-2);font-family:'Inter',sans-serif;font-size:12px;font-weight:500;cursor:pointer;transition:100ms}
  .btn:hover{background:var(--surface-hi);color:var(--text);border-color:var(--border-hi)}
  .btn.primary{background:var(--primary-hi);border-color:var(--primary-hi);color:#fff}
  .btn.primary:hover{background:var(--primary);border-color:var(--primary);color:#fff}
  .tc-grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(320px,1fr));gap:18px}
  .tc-card{position:relative;padding:22px 22px 20px;background:var(--surface);border:1px solid var(--border);border-radius:14px;cursor:pointer;transition:200ms cubic-bezier(0.16,1,0.3,1);overflow:hidden;display:flex;flex-direction:column;gap:14px;min-height:138px}
  .tc-card::before{content:'';position:absolute;left:0;top:0;bottom:0;width:3px;background:var(--accent,var(--primary));opacity:0.7;transition:200ms}
  .tc-card.type-Functional{--accent:var(--primary)}
  .tc-card.type-Negative{--accent:var(--rose)}
  .tc-card.type-Edge{--accent:var(--amber)}
  .tc-card.type-Integration{--accent:var(--violet)}
  .tc-card.type-Accessibility{--accent:var(--emerald)}
  .tc-card.type-Bug{--accent:var(--red)}
  .tc-card.type-Verification{--accent:var(--orange)}
  .tc-card:hover{transform:translateY(-3px);border-color:var(--border-hi);box-shadow:var(--shadow)}
  .tc-card:hover::before{opacity:1;width:4px}
  .tc-card .row1{display:flex;align-items:center;gap:10px;flex-wrap:wrap}
  .tc-card .tcid{font-family:'JetBrains Mono',monospace;font-size:11px;color:var(--text-3);font-weight:600;letter-spacing:0.05em}
  .tc-card h3{margin:0;font-size:15.5px;font-weight:700;color:var(--text);line-height:1.4;letter-spacing:-0.005em;overflow:hidden;display:-webkit-box;-webkit-line-clamp:2;-webkit-box-orient:vertical}
  .tc-card .meta-row{display:flex;align-items:center;gap:8px;flex-wrap:wrap}
  .tc-card .meta-row .tag{font-size:10.5px;font-weight:600;letter-spacing:0.04em;color:var(--text-3);text-transform:uppercase}
  .tc-card .meta-row .tag b{color:var(--text-2);font-weight:700;margin-left:3px;text-transform:none;letter-spacing:0}
  .tc-card .meta-row .tag.priority b{color:var(--text)}
  .tc-card .meta-row .tag.priority.Critical b{color:#b91c1c}
  .tc-card .meta-row .tag.priority.High b{color:#c2410c}
  .tc-card .meta-row .tag.priority.Medium b{color:#b45309}
  .tc-card .meta-row .tag.priority.Low b{color:#047857}
  .tc-card .meta-row .tag.priority.Serious b{color:#be123c}
  .tc-card .meta-row .tag.priority.Minor b{color:var(--slate)}
  .tc-card .meta-row .sep{color:var(--text-3);opacity:0.5}
  .tc-card .foot{display:flex;align-items:center;gap:8px;margin-top:auto;flex-wrap:wrap}
  /* status dot + label (subtle replacement for full pill on card) */
  .dot-status{display:inline-flex;align-items:center;gap:7px;font-size:12px;font-weight:600;color:var(--text-2)}
  .dot-status .dot{width:8px;height:8px;border-radius:50%;background:var(--text-3)}
  .dot-status[data-status="Pass"]{color:#047857}
  .dot-status[data-status="Pass"] .dot{background:var(--emerald);box-shadow:0 0 0 3px rgba(16,185,129,0.15)}
  .dot-status[data-status="Fail"]{color:#b91c1c}
  .dot-status[data-status="Fail"] .dot{background:var(--red);box-shadow:0 0 0 3px rgba(239,68,68,0.18)}
  .dot-status[data-status="On Hold"]{color:#b45309}
  .dot-status[data-status="On Hold"] .dot{background:var(--amber);box-shadow:0 0 0 3px rgba(245,158,11,0.18)}
  .pill{display:inline-flex;align-items:center;padding:2px 7px;font-size:10px;font-weight:700;text-transform:uppercase;letter-spacing:0.04em;border-radius:4px}
  .pill.type.Functional{background:var(--primary-soft);color:var(--primary-hi)}
  .pill.type.Negative{background:var(--rose-soft);color:#be123c}
  .pill.type.Edge{background:var(--amber-soft);color:#b45309}
  .pill.type.Integration{background:var(--violet-soft);color:#6d28d9}
  .pill.type.Accessibility{background:var(--emerald-soft);color:#047857}
  .pill.type.Bug{background:var(--red-soft);color:#b91c1c}
  .pill.type.Verification{background:var(--orange-soft);color:#c2410c}
  .pill.sev.Critical{background:var(--red-soft);color:#b91c1c}
  .pill.sev.High{background:var(--orange-soft);color:#c2410c}
  .pill.sev.Medium{background:var(--amber-soft);color:#b45309}
  .pill.sev.Low{background:var(--emerald-soft);color:#166534}
  .pill.sev.Serious{background:var(--rose-soft);color:#be123c}
  .pill.sev.Minor{background:var(--slate-soft);color:var(--slate)}
  .pill.sev.RESPONSIVE{background:var(--violet-soft);color:#6d28d9}
  .pill.status.Pass{background:var(--emerald-soft);color:#047857}
  .pill.status.Fail{background:var(--red-soft);color:#b91c1c}
  .pill.status.Open{background:var(--orange-soft);color:#c2410c}
  .pill.status[data-status="On Hold"]{background:var(--amber-soft);color:#b45309}
  .tc-card .roles{display:flex;gap:5px;flex-wrap:wrap}
  .role-badge{display:inline-flex;align-items:center;gap:4px;padding:2px 8px;font-size:10px;font-weight:600;background:var(--surface-2);color:var(--text-2);border:1px solid var(--border);border-radius:999px}
  .role-badge.role-Admin{color:var(--primary-hi);background:var(--primary-soft);border-color:var(--primary)}
  .role-badge.role-StateHead{color:var(--violet);background:var(--violet-soft);border-color:var(--violet)}
  .role-badge.role-GSCO{color:#047857;background:var(--emerald-soft);border-color:var(--emerald)}
  .role-badge.role-Any{color:var(--text-3)}
  .empty{grid-column:1/-1;padding:56px 20px;text-align:center;background:var(--surface);border:1px dashed var(--border);border-radius:12px;color:var(--text-3)}
  .empty b{display:block;color:var(--text);font-size:16px;font-weight:700;margin-bottom:4px}
  /* Side drawer (slides in from right) — replaces the old centered modal */
  .modal-bg{position:fixed;inset:0;background:rgba(15,23,42,0.45);backdrop-filter:blur(2px);-webkit-backdrop-filter:blur(2px);z-index:200;display:none;opacity:0;transition:opacity 220ms cubic-bezier(0.16,1,0.3,1)}
  .modal-bg.on{display:block;opacity:1}
  .modal{position:fixed;top:0;right:0;bottom:0;width:560px;max-width:100vw;background:var(--surface);box-shadow:-24px 0 60px rgba(15,23,42,0.20),-4px 0 12px rgba(15,23,42,0.08);overflow-y:auto;transform:translateX(100%);transition:transform 320ms cubic-bezier(0.16,1,0.3,1);padding:0}
  .modal-bg.on .modal{transform:translateX(0)}
  /* Sticky header inside drawer */
  .modal .drawer-head{position:sticky;top:0;background:var(--surface);padding:18px 28px 14px;border-bottom:1px solid var(--border);z-index:5;display:flex;align-items:center;gap:14px}
  .modal .drawer-head .tcid{font-family:'JetBrains Mono',monospace;color:var(--text-3);font-weight:600;font-size:11.5px;letter-spacing:0.06em}
  .modal .drawer-head .spacer{margin-left:auto}
  .modal .x{background:var(--surface-2);border:1px solid var(--border);width:32px;height:32px;border-radius:8px;cursor:pointer;color:var(--text-2);font-size:16px;line-height:1;transition:120ms;display:grid;place-items:center}
  .modal .x:hover{background:var(--surface-hi);color:var(--text);border-color:var(--border-hi)}
  .modal .drawer-body{padding:22px 28px 28px}
  .modal .modal-row1{display:flex;align-items:center;gap:10px;flex-wrap:wrap;margin-bottom:14px}
  .modal h2{margin:0 0 22px;font-size:22px;font-weight:700;letter-spacing:-0.012em;color:var(--text);line-height:1.3}
  .modal section{margin-top:24px;padding-top:0;border-top:none}
  .modal section:first-of-type{margin-top:0}
  .modal h4{font-size:10.5px;font-weight:700;color:var(--text-3);margin:0 0 10px;text-transform:uppercase;letter-spacing:0.12em}
  .modal pre{margin:0;font-family:'Inter',sans-serif;font-size:14px;line-height:1.65;color:var(--text-2);white-space:pre-wrap;word-wrap:break-word}
  .modal-actions{display:flex;gap:10px;align-items:center;flex-wrap:wrap;margin-top:28px;padding-top:22px;border-top:1px solid var(--border)}

  /* Touchpoints section — grouped checklist of role → screen → action */
  .touchpoints{display:flex;flex-direction:column;gap:14px}
  .touchpoints .role-group{padding:14px 16px;background:var(--surface-2);border:1px solid var(--border);border-radius:10px}
  .touchpoints .role-head{display:flex;align-items:center;gap:8px;font-size:12px;font-weight:700;color:var(--text);margin-bottom:10px;letter-spacing:0.01em}
  .touchpoints .role-head .role-badge{font-size:10px;padding:3px 9px}
  .touchpoints ul{margin:0;padding:0;list-style:none;display:flex;flex-direction:column;gap:8px}
  .touchpoints li{display:flex;gap:10px;align-items:flex-start;font-size:13.5px;color:var(--text-2);line-height:1.5}
  .touchpoints li::before{content:'';flex:0 0 6px;height:6px;border-radius:50%;background:var(--primary);margin-top:8px;opacity:0.45}
  .touchpoints .screen{font-weight:600;color:var(--text);margin-right:6px}
  .touchpoints .empty{color:var(--text-3);font-size:13px;font-style:italic;padding:10px 0}
  .status-edit{display:flex;align-items:center;gap:10px;flex-wrap:wrap}
  .status-edit label{font-size:12px;font-weight:600;color:var(--text-2)}
  .status-edit select{padding:7px 28px 7px 12px;border:1px solid var(--border);border-radius:8px;background-color:var(--surface);background-image:url("data:image/svg+xml;charset=utf-8,%3Csvg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 20 20' fill='%2364748b'%3E%3Cpath fill-rule='evenodd' d='M5.23 7.21a.75.75 0 011.06.02L10 11.06l3.71-3.83a.75.75 0 111.08 1.04l-4.25 4.39a.75.75 0 01-1.08 0L5.21 8.27a.75.75 0 01.02-1.06z' clip-rule='evenodd'/%3E%3C/svg%3E");background-repeat:no-repeat;background-position:right 10px center;background-size:14px;font-family:'Inter',sans-serif;font-size:13px;font-weight:600;cursor:pointer;appearance:none;-webkit-appearance:none;-moz-appearance:none;min-width:140px;outline:none}
  .status-edit select:focus{border-color:var(--primary);box-shadow:0 0 0 3px rgba(99,102,241,0.15)}
  .status-edit select[data-status="Pass"]{color:#047857;background-color:var(--emerald-soft);border-color:var(--emerald)}
  .status-edit select[data-status="Fail"]{color:#b91c1c;background-color:var(--red-soft);border-color:var(--red)}
  .status-edit select[data-status="On Hold"]{color:#b45309;background-color:var(--amber-soft);border-color:var(--amber)}
  .status-edit .save-ind{font-size:12px;color:var(--text-3);display:none}
  .status-edit .save-ind.on{display:inline}
  .status-edit .save-ind.ok{color:var(--emerald)}
  .status-edit .save-ind.err{color:var(--red)}
  @media (max-width:960px){.menu-btn{display:flex}.shell{grid-template-columns:1fr}.sidebar{position:fixed;top:var(--topbar-h);left:0;bottom:0;width:var(--sidebar-w);z-index:50;transform:translateX(-105%);transition:transform 220ms;box-shadow:var(--shadow-lg)}.sidebar.open{transform:translateX(0)}.topnav{display:none}.top-stats{display:none}main{padding:18px 14px 60px}}
  @media (max-width:520px){.tc-grid{grid-template-columns:1fr}.logo-sub{display:none}}

  /* JIRA ticket pill on TC cards (when a ticket already exists) */
  .pill.jira{background:rgba(99,102,241,0.15);color:#a5b4fc;border:1px solid rgba(165,180,252,0.35);font-family:'JetBrains Mono',monospace;font-size:10.5px}
  .pill.jira a{color:inherit;text-decoration:none}
  .pill.jira a:hover{text-decoration:underline}

  /* JIRA ticket form inside the TC modal */
  .jira-form{margin-top:18px;padding:18px;border-radius:12px;border:1px solid var(--border);background:rgba(99,102,241,0.04)}
  .jira-form h3{margin:0 0 4px;font-size:15px;display:flex;align-items:center;gap:8px}
  .jira-form .sub{color:var(--text-3);font-size:12.5px;margin-bottom:14px}
  .jira-form label{display:block;font-size:11.5px;letter-spacing:0.04em;text-transform:uppercase;color:var(--text-3);margin-bottom:6px;margin-top:12px;font-weight:600}
  .jira-form input[type=text],.jira-form textarea,.jira-form select{width:100%;background:var(--surface);border:1px solid var(--border);border-radius:8px;padding:9px 11px;color:var(--text);font:inherit;font-size:13.5px;box-sizing:border-box;transition:border-color 120ms}
  .jira-form input[type=text]:focus,.jira-form textarea:focus,.jira-form select:focus{outline:none;border-color:var(--primary)}
  .jira-form textarea{min-height:74px;resize:vertical;line-height:1.5;font-family:inherit}
  .jira-form .typetoggle{display:inline-flex;gap:0;background:var(--surface);border:1px solid var(--border);border-radius:8px;padding:3px;margin-bottom:4px}
  .jira-form .typetoggle button{background:transparent;border:0;color:var(--text-3);padding:5px 14px;border-radius:6px;cursor:pointer;font-size:12.5px;font-weight:600;transition:120ms}
  .jira-form .typetoggle button.on{background:var(--primary);color:white}
  .jira-form .dropzone{border:2px dashed var(--border-hi);border-radius:10px;padding:18px;text-align:center;color:var(--text-3);font-size:13px;cursor:pointer;transition:120ms;background:var(--surface)}
  .jira-form .dropzone.dragover{border-color:var(--primary);background:rgba(99,102,241,0.06);color:var(--primary-hi)}
  .jira-form .dropzone.has{border-style:solid;border-color:var(--emerald);color:var(--emerald);text-align:left;padding:10px 14px}
  .jira-form .dropzone img.thumb{max-width:100%;max-height:160px;border-radius:6px;display:block;margin-top:10px;border:1px solid var(--border)}
  .jira-form .preview-meta{display:flex;justify-content:space-between;align-items:center;gap:8px}
  .jira-form .dropzone .remove{background:transparent;border:0;color:var(--rose);cursor:pointer;font-size:12px}
  .jira-form .actions{margin-top:18px;display:flex;gap:8px;align-items:center;flex-wrap:wrap}
  .jira-form .status{font-size:12.5px;color:var(--text-3);margin-left:auto}
  .jira-form .status.ok{color:var(--emerald)}
  .jira-form .status.err{color:var(--rose)}
  .jira-form .result{margin-top:14px;padding:12px 14px;border-radius:8px;background:rgba(16,185,129,0.08);border:1px solid rgba(16,185,129,0.3);font-size:13px;display:none}
  .jira-form .result.on{display:block}
  .jira-form .result a{color:var(--emerald);font-weight:600;text-decoration:none;font-family:'JetBrains Mono',monospace}
  .jira-form .result a:hover{text-decoration:underline}
  .jira-form .open-btn{background:transparent;border:1px solid var(--border);color:var(--text-2);padding:6px 12px;border-radius:8px;cursor:pointer;font-size:12.5px;text-decoration:none;display:inline-flex;align-items:center;gap:6px;transition:120ms}
  .jira-form .open-btn:hover{border-color:var(--primary);color:var(--primary-hi)}
  .file-input-hidden{display:none}
</style>
</head>
<body>
<div class="topbar">
  <button class="menu-btn" id="menuBtn" aria-label="Menu">☰</button>
  <div class="logo" id="goHome">
    <div class="logo-mark">▲</div>
    <div>MEL Test Dashboard<div class="logo-sub" id="roleSummary"></div></div>
  </div>
  <div class="topnav" id="topnav"></div>
  <div class="top-stats" id="topStats"></div>
  <button class="logout-btn" id="logoutBtn" title="Sign out">⏻ Sign out</button>
</div>
<div class="shell">
  <aside class="sidebar" id="sidebar"></aside>
  <main id="root"></main>
</div>
<div class="modal-bg" id="modalBg"><div class="modal" id="modal"></div></div>

<script id="quest-data" type="application/json">__DATA__</script>
<script>
const DATA = JSON.parse(document.getElementById('quest-data').textContent);
const STATUS_OVERRIDES = {};   // tcId -> live status from /api/status (overrides embedded)
const JIRA_KEYS = {};          // tcId -> JIRA issue key from /api/jira/list
let ROLE = 'editor';           // 'editor' | 'viewer' — set from /api/whoami on load
const isViewer = () => ROLE === 'viewer';

function effectiveStatus(tc) {
  if (Object.prototype.hasOwnProperty.call(STATUS_OVERRIDES, tc.id)) {
    return STATUS_OVERRIDES[tc.id] || 'Not run';
  }
  return tc.status || 'Not run';
}

const state = {
  view: 'home', moduleId: null, role: 'All',
  filters: { type:'All', sev:'All', status:'All' },
  search: '',
};

const TYPES = ['Functional','Negative','Edge','Integration','Accessibility','Bug','Verification'];
const SEVS  = ['Critical','High','Medium','Low','Serious','Minor'];
const STATS = ['Not run','Pass','Fail','On Hold','Open'];
const ROLES = ['Admin','State Head','GSCO'];
const STATUS_CHOICES = ['Not run','Pass','Fail','On Hold'];

function el(tag, attrs={}, ...children) {
  const e = document.createElement(tag);
  for (const k in attrs) {
    if (k === 'class') e.className = attrs[k];
    else if (k === 'on') for (const ev in attrs.on) e.addEventListener(ev, attrs.on[ev]);
    else if (k === 'html') e.innerHTML = attrs[k];
    else e.setAttribute(k, attrs[k]);
  }
  for (const c of children) {
    if (c == null) continue;
    if (typeof c === 'string' || typeof c === 'number') e.appendChild(document.createTextNode(c));
    else e.appendChild(c);
  }
  return e;
}
function tcMatchesRole(tc) {
  if (state.role === 'All') return true;
  return tc.roles.includes(state.role);
}

async function loadRole() {
  try {
    const r = await fetch('/api/whoami', { credentials: 'same-origin' });
    if (!r.ok) return;
    const data = await r.json();
    if (data && data.role) ROLE = data.role;
  } catch (e) { /* default to editor */ }
}

async function loadJiraKeys() {
  try {
    const r = await fetch('/api/jira/list', { credentials: 'same-origin' });
    if (!r.ok) return;
    const data = await r.json();
    for (const k in data) JIRA_KEYS[k] = data[k];
  } catch (e) { /* offline fallback */ }
}

async function loadStatusOverrides() {
  try {
    const r = await fetch('/api/status', { credentials:'same-origin' });
    if (!r.ok) return;
    const data = await r.json();
    for (const k in data) STATUS_OVERRIDES[k] = data[k];
  } catch (e) { /* offline fallback to embedded statuses */ }
}

async function setStatusOnServer(tcId, status) {
  const r = await fetch('/api/status', {
    method: 'POST',
    headers: { 'content-type': 'application/json' },
    credentials: 'same-origin',
    body: JSON.stringify({ tcId, status }),
  });
  if (!r.ok) throw new Error(r.status);
  if (!status || status === 'Not run') {
    delete STATUS_OVERRIDES[tcId];
  } else {
    STATUS_OVERRIDES[tcId] = status;
  }
}

function renderTopNav() {
  const top = document.getElementById('topnav');
  top.innerHTML = '';
  if (isViewer()) {
    top.appendChild(el('span',{class:'role-chip',style:'background:rgba(245,158,11,0.15);color:#fbbf24;border-color:rgba(251,191,36,0.35);cursor:default;'}, '👁 Read-only'));
  }
  for (const r of ['All','Admin','State Head','GSCO']) {
    top.appendChild(el('span', {
      class:'role-chip' + (state.role===r?' on':''),
      'data-role': r,
      on: { click: () => { state.role = r; render(); } }
    }, r==='All' ? null : el('span',{class:'dot'}), r));
  }
}
function renderTopStats() {
  const stats = document.getElementById('topStats');
  const sub = document.getElementById('roleSummary');
  let total=0,pass=0,fail=0,hold=0,notRun=0;
  for (const m of DATA.modules) for (const tc of m.tcs) {
    if (!tcMatchesRole(tc)) continue;
    total++;
    const s = effectiveStatus(tc);
    if (s==='Pass') pass++;
    else if (s==='Fail') fail++;
    else if (s==='On Hold') hold++;
    else if (s==='Not run' || s==='') notRun++;
  }
  stats.innerHTML = '';
  stats.appendChild(el('span',{class:'stat-pill'},el('b',{},String(total)),' tests'));
  stats.appendChild(el('span',{class:'stat-pill pass'},el('b',{},String(pass)),' pass'));
  stats.appendChild(el('span',{class:'stat-pill fail'},el('b',{},String(fail)),' fail'));
  stats.appendChild(el('span',{class:'stat-pill hold'},el('b',{},String(hold)),' on hold'));
  stats.appendChild(el('span',{class:'stat-pill idle'},el('b',{},String(notRun)),' not run'));
  stats.appendChild(el('span',{class:'top-meta'},'· ' + DATA.generatedAt));
  sub.textContent = state.role==='All' ? 'All roles · ' + DATA.totalTcs + ' total tests' : state.role + ' · ' + total + ' tests';
}
function renderSidebar() {
  const sb = document.getElementById('sidebar');
  sb.innerHTML = '';
  sb.appendChild(el('div',{class:'home-link'+(state.view==='home'?' active':''),on:{click:goHome}},'⌂ ','Dashboard Home'));
  const byCat = {};
  for (const m of DATA.modules) (byCat[m.category] ||= []).push(m);
  for (const cat in byCat) {
    sb.appendChild(el('div',{class:'category'}, cat));
    for (const m of byCat[cat]) {
      const visible = m.tcs.filter(tcMatchesRole).length;
      sb.appendChild(el('div',{
        class:'module-link'+(state.moduleId===m.id?' active':''),
        on:{click: () => openModule(m.id)}
      }, el('span',{class:'ic'},m.icon), el('span',{}, m.name), el('span',{class:'count'},String(visible))));
    }
  }
}
function goHome() {
  state.view='home'; state.moduleId=null;
  state.filters={type:'All',sev:'All',status:'All'}; state.search='';
  closeSidebar(); render();
}
function openModule(id) {
  state.view='module'; state.moduleId=id;
  state.filters={type:'All',sev:'All',status:'All'}; state.search='';
  closeSidebar(); render();
  window.scrollTo({top:0});
}
function renderHome() {
  const root = document.getElementById('root');
  root.innerHTML = '';
  const hero = el('div',{class:'hero'},
    el('h1',{},'Test Catalogue'),
    el('p',{},`Browse ${DATA.totalTcs} test cases across ${DATA.modules.length} modules. Use the role buttons on top to filter; pick a module from the sidebar to dig in.`)
  );
  if (state.role !== 'All') {
    const c = DATA.modules.reduce((acc,m)=>acc+m.tcs.filter(tcMatchesRole).length, 0);
    hero.appendChild(el('div',{class:'role-banner'},'Filtering by ',el('b',{},state.role),` · ${c} tests match`));
  }
  root.appendChild(hero);
  const byCat = {};
  for (const m of DATA.modules) (byCat[m.category] ||= []).push(m);
  for (const cat in byCat) {
    const block = el('div',{class:'cat-block'},
      el('div',{class:'cat-title'},cat),
      el('div',{class:'module-grid'}, ...byCat[cat].map(m=>moduleCard(m)))
    );
    root.appendChild(block);
  }
}
function moduleCard(m) {
  const visible = m.tcs.filter(tcMatchesRole);
  const total = visible.length;
  const pass = visible.filter(t => effectiveStatus(t)==='Pass').length;
  const fail = visible.filter(t => effectiveStatus(t)==='Fail').length;
  const tested = pass + fail;
  const pct = total ? Math.round(tested/total*100) : 0;
  return el('div',{class:'module-card', on:{click:()=>openModule(m.id)}},
    el('span',{class:'ic'}, m.icon),
    el('div',{class:'name'}, m.name),
    el('div',{class:'meta'}, String(total)+' tests · '+String(tested)+' covered'),
    el('div',{class:'progress'}, el('div',{style:`width:${pct}%`}))
  );
}
function renderModule() {
  const root = document.getElementById('root');
  root.innerHTML = '';
  const m = DATA.modules.find(x=>x.id===state.moduleId);
  if (!m) { goHome(); return; }
  root.appendChild(el('div',{class:'crumb'},
    el('a',{on:{click:goHome}},'⌂ Home'), el('span',{},' › '),
    el('span',{},m.category), el('span',{},' › '), el('span',{},m.name)));
  root.appendChild(el('div',{class:'page-head'},
    el('div',{class:'ic'},m.icon),
    el('div',{}, el('h1',{},m.name),
      el('div',{class:'sub'}, `${m.tcs.length} test cases · ${state.role==='All'?'All roles':state.role+' filter'}`))));
  const fb = el('div',{class:'filter-bar'});
  fb.appendChild(makeSelect('Type','type',['All',...TYPES]));
  fb.appendChild(makeSelect('Severity','sev',['All',...SEVS]));
  fb.appendChild(makeSelect('Status','status',['All',...STATS]));
  const search = el('input',{class:'search',type:'search',placeholder:'Search id, scenario, steps, expected…'});
  search.value = state.search;
  search.addEventListener('input', e => { state.search = e.target.value; renderGrid(m); });
  fb.appendChild(search);
  fb.appendChild(el('button',{class:'btn', on:{click:()=>exportCsv(m)}}, '⤓ Export CSV'));
  root.appendChild(fb);
  root.appendChild(el('div',{class:'tc-grid',id:'tcGrid'}));
  renderGrid(m);
}
function makeSelect(label,key,options) {
  const isActive = state.filters[key] && state.filters[key]!=='All';
  const sel = el('select',{class:'fselect'+(isActive?' active':''),'aria-label':label,
    on:{change: e => { state.filters[key]=e.target.value; renderModule(); }}});
  for (const opt of options) {
    const o = document.createElement('option');
    o.value = opt;
    o.textContent = opt==='All' ? label+': All' : opt;
    if (state.filters[key]===opt) o.selected = true;
    sel.appendChild(o);
  }
  return sel;
}
function filteredTcs(m) {
  const f = state.filters;
  const q = state.search.trim().toLowerCase();
  return m.tcs.filter(tc => {
    if (!tcMatchesRole(tc)) return false;
    if (f.type!=='All' && tc.type!==f.type) return false;
    if (f.sev!=='All' && tc.severity!==f.sev) return false;
    if (f.status!=='All' && effectiveStatus(tc)!==f.status) return false;
    if (q) {
      const blob = [tc.id, tc.scenario, tc.pre, tc.steps, tc.expected, tc.notes].join(' ').toLowerCase();
      if (!blob.includes(q)) return false;
    }
    return true;
  });
}
function renderGrid(m) {
  const grid = document.getElementById('tcGrid');
  grid.innerHTML = '';
  const tcs = filteredTcs(m);
  if (!tcs.length) {
    grid.appendChild(el('div',{class:'empty'}, el('b',{},'No tests match the filters'), 'Try clearing the role, type, severity, status filters, or the search.'));
    return;
  }
  for (const tc of tcs) grid.appendChild(tcCard(tc));
}
function tcCard(tc) {
  const c = el('div',{class:'tc-card type-'+(tc.type||'Functional'), on:{click:()=>openTc(tc)}});
  const s = effectiveStatus(tc);
  const jiraKey = JIRA_KEYS[tc.id];
  c.appendChild(el('div',{class:'row1'},
    el('span',{class:'tcid'}, tc.id)
  ));
  c.appendChild(el('h3',{}, tc.scenario || '(no name)'));
  // Type + Priority labels (compact, scannable, not loud pills)
  const meta = el('div',{class:'meta-row'});
  meta.appendChild(el('span',{class:'tag'}, 'Type', el('b',{}, ' ', tc.type || 'Functional')));
  meta.appendChild(el('span',{class:'sep'}, '·'));
  meta.appendChild(el('span',{class:'tag priority '+(tc.severity||'')}, 'Priority', el('b',{}, ' ', tc.severity || '—')));
  c.appendChild(meta);
  const foot = el('div',{class:'foot'});
  if (s && s!=='Not run') {
    foot.appendChild(el('span',{class:'dot-status','data-status':s}, el('span',{class:'dot'}), s));
  } else {
    foot.appendChild(el('span',{class:'dot-status'}, el('span',{class:'dot'}), 'Not run'));
  }
  if (jiraKey) foot.appendChild(el('span',{class:'pill jira',style:'margin-left:auto;', title:'JIRA ticket'}, '🎫 ', jiraKey));
  c.appendChild(foot);
  return c;
}
function openTc(tc) {
  const drawer = document.getElementById('modal');
  drawer.innerHTML = '';
  const s = effectiveStatus(tc);
  const jiraKey = JIRA_KEYS[tc.id];
  // Sticky drawer header (TC ID + close button)
  drawer.appendChild(el('div',{class:'drawer-head'},
    el('span',{class:'tcid'}, tc.id),
    el('span',{class:'spacer'}),
    el('button',{class:'x','aria-label':'Close',on:{click:closeTc}}, '×')
  ));
  // Scrollable body — alias `m` so existing appendChild calls below target the body
  const m = el('div',{class:'drawer-body'});
  drawer.appendChild(m);
  m.appendChild(el('div',{class:'modal-row1'},
    el('span',{class:'pill type '+(tc.type||'')}, tc.type||'Functional'),
    el('span',{class:'pill sev '+(tc.severity||'')}, tc.severity||'—'),
    (s && s!=='Not run') ? el('span',{class:'pill status '+s,'data-status':s}, s) : null,
    jiraKey ? el('span',{class:'pill jira'},
      el('a',{href:`https://enableindiaorg.atlassian.net/browse/${jiraKey}`,target:'_blank',rel:'noopener'}, '🎫 ', jiraKey)) : null
  ));
  m.appendChild(el('h2',{}, tc.scenario || '(no name)'));
  if (tc.roles && tc.roles.length) {
    const r = el('div',{class:'roles', style:'margin-bottom:18px;'});
    for (const role of tc.roles) {
      const cls = role==='State Head' ? 'role-StateHead' : 'role-'+role.replace(/\s/g,'');
      r.appendChild(el('span',{class:'role-badge '+cls},
        role==='Admin'?'👑 Admin':role==='State Head'?'🛡 State Head':role==='GSCO'?'🎯 GSCO':'◇ '+role));
    }
    m.appendChild(r);
  }
  // Status edit row
  const sEdit = el('div',{class:'status-edit'});
  sEdit.appendChild(el('label',{},'Status:'));
  const sel = el('select',{'data-status':s,'aria-label':'Status'});
  if (isViewer()) sel.disabled = true;
  for (const opt of STATUS_CHOICES) {
    const o = document.createElement('option');
    o.value = opt; o.textContent = opt;
    if (s===opt) o.selected = true;
    sel.appendChild(o);
  }
  const ind = el('span',{class:'save-ind'},'');
  sel.addEventListener('change', async () => {
    const newStatus = sel.value;
    sel.setAttribute('data-status', newStatus);
    ind.className = 'save-ind on'; ind.textContent = 'Saving…';
    try {
      await setStatusOnServer(tc.id, newStatus);
      ind.className = 'save-ind on ok'; ind.textContent = '✓ Saved';
      renderTopStats(); renderSidebar();
      // Refresh card if module view is active
      if (state.view==='module') {
        const mod = DATA.modules.find(x=>x.id===state.moduleId);
        if (mod) renderGrid(mod);
      } else {
        renderHome();
      }
      setTimeout(()=>{ ind.className='save-ind'; ind.textContent=''; }, 1600);
    } catch (e) {
      ind.className = 'save-ind on err'; ind.textContent = '✕ Save failed';
    }
  });
  sEdit.appendChild(sel); sEdit.appendChild(ind);
  m.appendChild(el('section',{},
    el('h4',{},'Test Status'),
    sEdit
  ));
  if (tc.pre)      m.appendChild(section('Precondition', tc.pre));
  if (tc.steps)    m.appendChild(section('Steps', tc.steps));
  if (tc.expected) m.appendChild(section('Expected Result', tc.expected));
  m.appendChild(renderTouchpoints(tc));
  if (tc.notes)    m.appendChild(section('Notes', tc.notes));
  if (tc.br)       m.appendChild(section('References', tc.br));
  if (!isViewer()) {
    const actions = el('div',{class:'modal-actions'});
    actions.appendChild(el('button',{class:'btn primary',on:{click:()=>renderJiraForm(tc)}}, JIRA_KEYS[tc.id] ? '🎫 File another ticket' : '🎫 Create JIRA ticket'));
    m.appendChild(actions);
  }
  // Anchor where the JIRA ticket form will mount on demand
  m.appendChild(el('div',{id:'jiraMount'}));
  document.getElementById('modalBg').classList.add('on');
}
function renderJiraForm(tc) {
  const mount = document.getElementById('jiraMount');
  if (!mount) return;
  mount.innerHTML = '';
  const form = el('div',{class:'jira-form'});
  form.appendChild(el('h3',{}, '🎫 New JIRA ticket', el('span',{style:'font-weight:400;color:var(--text-3);font-size:13px;'},' · MPSW')));
  form.appendChild(el('div',{class:'sub'}, `Filed as ${tc.id} · reporter: dev.csv@enableindia.org · default unassigned`));

  // Issue type toggle
  form.appendChild(el('label',{},'Issue type'));
  const typeWrap = el('div',{class:'typetoggle'});
  const tBug = el('button',{class:'on',type:'button'},'Bug');
  const tTask = el('button',{type:'button'},'Task');
  let selectedType = 'Bug';
  tBug.addEventListener('click', () => { selectedType='Bug'; tBug.classList.add('on'); tTask.classList.remove('on'); });
  tTask.addEventListener('click', () => { selectedType='Task'; tTask.classList.add('on'); tBug.classList.remove('on'); });
  typeWrap.appendChild(tBug); typeWrap.appendChild(tTask);
  form.appendChild(typeWrap);

  // Title
  form.appendChild(el('label',{},'Title'));
  const titleInput = el('input',{type:'text', value: tc.scenario || ''});
  form.appendChild(titleInput);

  // Steps
  form.appendChild(el('label',{},'Steps to reproduce'));
  const stepsInput = el('textarea',{}, tc.steps || '');
  form.appendChild(stepsInput);

  // Expected
  form.appendChild(el('label',{},'Expected result'));
  const expInput = el('textarea',{}, tc.expected || '');
  form.appendChild(expInput);

  // Actual
  form.appendChild(el('label',{},'Actual result'));
  const actInput = el('textarea',{placeholder:'What actually happened? Include error messages, behaviour observed, etc.'},'');
  form.appendChild(actInput);

  // Screenshot
  form.appendChild(el('label',{},'Screenshot (optional)'));
  const drop = el('div',{class:'dropzone',tabIndex:0,role:'button','aria-label':'Upload screenshot'});
  drop.innerHTML = '<div>📎 Drop an image here, paste from clipboard, or click to select</div>';
  const fileIn = el('input',{type:'file',accept:'image/*',class:'file-input-hidden'});
  drop.appendChild(fileIn);
  let screenshotData = null;
  let screenshotName = null;
  function setFile(file) {
    if (!file || !file.type.startsWith('image/')) return;
    const reader = new FileReader();
    reader.onload = () => {
      const b64 = String(reader.result || '').split(',')[1] || '';
      screenshotData = b64;
      screenshotName = file.name || 'screenshot.png';
      drop.classList.add('has');
      drop.innerHTML = '';
      const meta = el('div',{class:'preview-meta'},
        el('span',{}, '✓ ', screenshotName, ' (', Math.round(file.size/1024), ' KB)'),
        el('button',{class:'remove',type:'button',on:{click:e=>{e.stopPropagation(); resetDrop();}}}, '× Remove'));
      drop.appendChild(meta);
      drop.appendChild(el('img',{class:'thumb',src:reader.result}));
    };
    reader.readAsDataURL(file);
  }
  function resetDrop() {
    screenshotData = null; screenshotName = null;
    drop.classList.remove('has','dragover');
    drop.innerHTML = '<div>📎 Drop an image here, paste from clipboard, or click to select</div>';
    drop.appendChild(fileIn);
  }
  drop.addEventListener('click', () => fileIn.click());
  drop.addEventListener('keydown', e => { if (e.key==='Enter'||e.key===' '){e.preventDefault();fileIn.click();} });
  fileIn.addEventListener('change', () => { if (fileIn.files && fileIn.files[0]) setFile(fileIn.files[0]); });
  drop.addEventListener('dragover', e => { e.preventDefault(); drop.classList.add('dragover'); });
  drop.addEventListener('dragleave', () => drop.classList.remove('dragover'));
  drop.addEventListener('drop', e => {
    e.preventDefault(); drop.classList.remove('dragover');
    if (e.dataTransfer && e.dataTransfer.files && e.dataTransfer.files[0]) setFile(e.dataTransfer.files[0]);
  });
  // Paste from clipboard while form has focus
  const pasteHandler = e => {
    if (!form.contains(document.activeElement) && document.activeElement !== form) return;
    const items = (e.clipboardData || {}).items || [];
    for (const it of items) if (it.type && it.type.startsWith('image/')) { setFile(it.getAsFile()); e.preventDefault(); return; }
  };
  document.addEventListener('paste', pasteHandler);
  form.appendChild(drop);

  // Actions
  const actions = el('div',{class:'actions'});
  const createBtn = el('button',{class:'btn primary',type:'button'}, 'Create ticket');
  const cancelBtn = el('button',{class:'btn',type:'button',on:{click:()=>{document.removeEventListener('paste',pasteHandler); mount.innerHTML='';}}}, 'Cancel');
  const status = el('span',{class:'status'},'');
  actions.appendChild(createBtn);
  actions.appendChild(cancelBtn);
  actions.appendChild(status);
  form.appendChild(actions);

  const result = el('div',{class:'result'});
  form.appendChild(result);

  createBtn.addEventListener('click', async () => {
    const title = titleInput.value.trim();
    if (!title) { status.className='status err'; status.textContent='Title required'; return; }
    createBtn.disabled = true; cancelBtn.disabled = true;
    status.className = 'status'; status.textContent = 'Filing in JIRA…';
    try {
      const payload = {
        tcId: tc.id,
        issueType: selectedType,
        title,
        steps: stepsInput.value,
        expected: expInput.value,
        actual: actInput.value,
        severity: tc.severity || '',
      };
      if (screenshotData) {
        payload.screenshotBase64 = screenshotData;
        payload.screenshotName = screenshotName;
      }
      const r = await fetch('/api/jira/create', {
        method: 'POST',
        headers: { 'content-type':'application/json' },
        credentials: 'same-origin',
        body: JSON.stringify(payload),
      });
      const data = await r.json();
      if (!r.ok) {
        status.className = 'status err';
        status.textContent = 'Failed: ' + (data && (data.error || JSON.stringify(data)) || r.status);
        createBtn.disabled = false; cancelBtn.disabled = false;
        return;
      }
      JIRA_KEYS[tc.id] = data.key;
      status.className = 'status ok'; status.textContent = '✓ Filed';
      result.classList.add('on');
      result.innerHTML = '';
      result.appendChild(el('div',{},
        '✓ Ticket created: ',
        el('a',{href:data.url,target:'_blank',rel:'noopener'}, data.key),
        data.attachment ? (data.attachment.ok ? ' · screenshot attached' : ' · ⚠ attachment failed') : ''));
      // Re-render the grid / home to reflect the new badge
      if (state.view==='module') {
        const mod = DATA.modules.find(x=>x.id===state.moduleId);
        if (mod) renderGrid(mod);
      } else { renderHome(); }
      // Keep form visible (user might want to file another); reset Actual + screenshot
      actInput.value = '';
      resetDrop();
      createBtn.disabled = false; cancelBtn.disabled = false;
    } catch (e) {
      status.className = 'status err'; status.textContent = 'Error: ' + (e?.message || e);
      createBtn.disabled = false; cancelBtn.disabled = false;
    }
  });

  mount.appendChild(form);
  // Cleanup paste handler when modal closes
  const cleanup = () => { document.removeEventListener('paste', pasteHandler); document.getElementById('modalBg').removeEventListener('transitionend', cleanup); };
  document.getElementById('modalBg').addEventListener('transitionend', cleanup, {once:true});
  titleInput.focus();
}
function section(label,text){return el('section',{},el('h4',{},label),el('pre',{},text));}
function renderTouchpoints(tc) {
  const sec = el('section',{});
  sec.appendChild(el('h4',{}, 'Touchpoints to verify'));
  const list = tc.touchpoints || [];
  if (!list.length) {
    sec.appendChild(el('div',{class:'empty', style:'color:var(--text-3);font-size:13px;font-style:italic;padding:6px 0 0;'},
      'No touchpoints recorded yet — the developer should verify only the primary screen for this test.'));
    return sec;
  }
  // Group by role; preserve first-seen order
  const byRole = new Map();
  for (const tp of list) {
    const r = tp.role || 'Any';
    if (!byRole.has(r)) byRole.set(r, []);
    byRole.get(r).push(tp);
  }
  const wrap = el('div',{class:'touchpoints'});
  for (const [role, items] of byRole.entries()) {
    const grp = el('div',{class:'role-group'});
    const cls = role==='State Head' ? 'role-StateHead' : 'role-'+role.replace(/\s/g,'');
    grp.appendChild(el('div',{class:'role-head'},
      el('span',{class:'role-badge '+cls},
        role==='Admin'?'👑 Admin':role==='State Head'?'🛡 State Head':role==='GSCO'?'🎯 GSCO':'◇ '+role)
    ));
    const ul = el('ul',{});
    for (const it of items) {
      const li = el('li',{});
      if (it.screen) li.appendChild(el('span',{class:'screen'}, it.screen, ' — '));
      li.appendChild(document.createTextNode(it.action || ''));
      ul.appendChild(li);
    }
    grp.appendChild(ul);
    wrap.appendChild(grp);
  }
  sec.appendChild(wrap);
  return sec;
}
function closeTc(){document.getElementById('modalBg').classList.remove('on');}
document.getElementById('modalBg').addEventListener('click',e=>{if(e.target===document.getElementById('modalBg'))closeTc();});
document.addEventListener('keydown',e=>{if(e.key==='Escape')closeTc();});
document.getElementById('goHome').addEventListener('click',goHome);
document.getElementById('logoutBtn').addEventListener('click', async () => {
  try { await fetch('/api/logout',{method:'POST',credentials:'same-origin'}); } catch {}
  location.replace('/login.html');
});
function copyTc(btn,tc,fmt) {
  const s = effectiveStatus(tc);
  let text;
  if (fmt==='json') {
    text = JSON.stringify({...tc, status: s}, null, 2);
  } else {
    text = `# ${tc.id} — ${tc.scenario}\n\n` +
      `**Type:** ${tc.type}  ·  **Severity:** ${tc.severity}  ·  **Status:** ${s}\n` +
      `**Roles:** ${(tc.roles||[]).join(', ')}\n\n` +
      (tc.pre      ? `## Precondition\n${tc.pre}\n\n` : '') +
      (tc.steps    ? `## Steps\n${tc.steps}\n\n` : '') +
      (tc.expected ? `## Expected\n${tc.expected}\n\n` : '') +
      (tc.notes    ? `## Notes\n${tc.notes}\n\n` : '') +
      (tc.br       ? `## References\n${tc.br}\n` : '');
  }
  navigator.clipboard.writeText(text).then(() => {
    const orig = btn.textContent;
    btn.textContent = '✓ Copied';
    setTimeout(()=>{ btn.textContent = orig; }, 1300);
  });
}
function exportCsv(m) {
  const rows = filteredTcs(m);
  if (!rows.length) { alert('Nothing to export — clear some filters first.'); return; }
  const headers = ['TC ID','Scenario','Type','Severity','Status','Roles','Precondition','Steps','Expected Result','References','Notes'];
  const csv = [headers.join(',')];
  for (const r of rows) {
    const cells = [r.id, r.scenario, r.type, r.severity, effectiveStatus(r),
      (r.roles||[]).join('; '), r.pre, r.steps, r.expected, r.br, r.notes];
    csv.push(cells.map(csvEscape).join(','));
  }
  const blob = new Blob(['﻿' + csv.join('\n')], {type:'text/csv;charset=utf-8'});
  const url = URL.createObjectURL(blob);
  const a = document.createElement('a');
  a.href = url; a.download = `${m.id}-filtered.csv`;
  document.body.appendChild(a); a.click(); document.body.removeChild(a);
  setTimeout(()=>URL.revokeObjectURL(url), 1500);
}
function csvEscape(v) {
  v = (v ?? '').toString();
  if (/[",\n]/.test(v)) return '"' + v.replace(/"/g,'""') + '"';
  return v;
}
function closeSidebar() { document.getElementById('sidebar').classList.remove('open'); }
document.getElementById('menuBtn').addEventListener('click', () => {
  document.getElementById('sidebar').classList.toggle('open');
});
function render() {
  renderTopNav(); renderTopStats(); renderSidebar();
  if (state.view==='home') renderHome(); else renderModule();
}
Promise.allSettled([loadRole(), loadStatusOverrides(), loadJiraKeys()]).then(render).catch(render);
</script>
</body>
</html>
"""

html = HTML.replace('__DATA__', DATA_JSON)
OUT.write_text(html, encoding='utf-8')
print(f'Generated: {OUT}')
print(f'Total modules: {len(modules_out)}')
print(f'Total TCs:     {total_tcs}')
print(f'File size:     {OUT.stat().st_size/1024:.1f} KB')
